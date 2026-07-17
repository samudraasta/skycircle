import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

MASTER_MENTOR = {
    'BC X-1': 'Rony', 'BC X-2': 'Joko', 'BC X-3': 'Angga', 'BC X-4': 'Arumbinang', 
    'BC X-5': 'Topik & Adrian', 'BC X-6': 'Rizki Putra', 'BC X-7': 'Rizza Al-Farisi', 
    'BC X-8': 'Teguh', 'BC X-9': 'Veril (XII)', 'BC X-10': 'Irghi (XII)', 
    'BC X-11': 'Khadafi (XII)', 'BC X-12': 'Almeer (XII)', 'SUMEN BC': 'Rekhsa',
    'GC1 X-1': 'Syifa', 'GC1 X-2': 'Rania (XII)', 'GC1 X-3': 'Hanin', 'GC1 X-4': 'Najla',
    'GC1 X-5': 'Hana', 'GC1 X-6': 'Khairunnisa (XII)', 'GC1 X-7': 'Shafa', 'GC1 X-8': 'Sakina',
    'GC1 X-9': 'Tiara (XII)', 'GC1 X-10': 'Citra (XII)', 'GC1 X-11': 'Nia', 'GC1 X-12': 'Ninda (XII)',
    'SUMEN1 GC': 'Ayun', 'SUMEN2 GC': 'Iqlima',
    'GC2 X-1': 'Oshe (XI)', 'GC2 X-2': 'Nadine (XI)', 'GC2 X-3': 'Zhafira (XI)', 'GC2 X-4': 'Faliqa (XI)',
    'GC2 X-5': 'Sabika (XI)', 'GC2 X-6': 'Bumi (XI)', 'GC2 X-7': 'Desya (XI)', 'GC2 X-8': 'Arsyila (XI)',
    'GC2 X-9': 'Almira (XI)', 'GC2 X-10': 'Lexa (XI)', 'GC2 X-11': 'Annisa Alya (XII)', 'GC2 X-12': 'Fiona (XI)'
}

def is_nonis(name):
    name = str(name).lower()
    muslim_keywords = ['muhammad', 'ahmad', 'achmad', 'farhan', 'firza', 'bilqis', 'nasyifa', 'zahra', 'hafizh', 'musyaffa', 'taufik', 'hidayah', 'abdillah', 'rafa', 'rizzieq', 'yusuf', 'faiha', 'raqilla', 'latief', 'benzema']
    nonis_keywords = ['christ', 'johan', 'ruth', 'floren', 'gabriel', 'samuel', 'gracia', 'maria', 'daniel', 'theresia', 'agatha', 'putu', 'harefa', 'simangunsong', 'sitorus', 'sitanggang', 'harianja', 'sirait', 'manurung', 'siahaan', 'sianipar', 'torus', 'ornando', 'hia', 'azalia', 'claudia', 'jesslyn', 'kariswan', 'melani', 'lionel', 'oryza', 'fricilia', 'keynia', 'laura', 'priscilia', 'bernala', 'defrino', 'joceline', 'quinsha', 'raynar', 'enzo', 'jeremia', 'nava', 'vincentius', 'chntya', 'henokh', 'weny', 'yoanne', 'avemarianne', 'olivia', 'lubis']
    
    for k in muslim_keywords:
        if k in name: return False
    for k in nonis_keywords:
        if k in name: return True
        
    return True # Default to Nonis if they don't match Muslim keywords and are unmapped

def generate_rekap():
    input_file = "./# Presensi SCC/1. Presensi SCC.xlsx"
    smavo_file = "./# Presensi SCC/SMAVO ANGKATAN 33 +SUPERMENTORING (2025-2026).xlsx"
    output_file = "Rekap_Presensi_SCC_X_XI_XII.xlsx"
    
    print(f"Membaca data master dari {input_file}...")
    try:
        df_master = pd.read_excel(input_file, sheet_name='Database Presensi SCC')
        
        try:
            df_31 = pd.read_excel(input_file, sheet_name='Data Siswa Mentah 31 (Cleaning)', header=None)
            df_31_clean = df_31.dropna(subset=[1, 4])
            df_31_clean = df_31_clean[df_31_clean[1].apply(lambda x: str(x).isdigit())]
            
            list_31 = []
            for _, row in df_31_clean.iterrows():
                list_31.append({
                    'NAMA': str(row[4]),
                    'NISN / NIS': f"{row[3]} / {row[2]}",
                    'L/P': str(row[5]),
                    'ANGKATAN': 31,
                    'KELAS TA': 'XII',
                    'Grup TA': '',
                    'Mentor TA': ''
                })
            df_master_31 = pd.DataFrame(list_31)
            df_master = pd.concat([df_master, df_master_31], ignore_index=True)
            print("Berhasil menggabungkan Angkatan 31!")
        except Exception as e:
            print(f"Gagal mengambil Angkatan 31: {e}")

        # Map SMAVO
        mentor_map = {}
        try:
            print(f"Membaca data Grup dan Mentor Angkatan 33 dari {smavo_file}...")
            df_smavo = pd.read_excel(smavo_file, sheet_name='Angkatan 33', header=1, dtype={'NISN': str})
            for _, row in df_smavo.iterrows():
                nisn_raw = str(row.get('NISN', '')).strip()
                nisn_clean = nisn_raw.lstrip('0')
                
                grup = str(row.get('Grup', '')).strip()
                mentor = str(row.get('Mentor', '')).strip()
                
                if grup.lower() == 'nan': grup = ''
                if mentor.lower() == 'nan': mentor = ''
                
                if nisn_clean:
                    mentor_map[nisn_clean] = {'Grup': grup, 'Mentor': mentor}
            print("Berhasil memetakan Grup & Mentor Angkatan 33!")
        except Exception as e:
            print(f"Gagal mengambil data SMAVO: {e}")

        # Pre-process girls for GC1/GC2 auto-assign
        print("Mempersiapkan auto-assign untuk siswi Perempuan...")
        df_33_girls = df_master[(df_master['ANGKATAN'] == 33) & (df_master['L/P'].str.strip().str.upper() == 'P')].copy()
        df_33_girls['is_nonis'] = df_33_girls['NAMA'].apply(lambda x: is_nonis(x))
        df_33_girls = df_33_girls[~df_33_girls['is_nonis']] # only muslim girls
        df_33_girls = df_33_girls.sort_values(by=['KELAS TA', 'NAMA'])

        girl_group_map = {}
        for kelas, group in df_33_girls.groupby('KELAS TA'):
            kelas_str = str(kelas).strip()
            kelas_fmt = kelas_str.replace('X', 'X-') if kelas_str.startswith('X') and '-' not in kelas_str else kelas_str
            
            n = len(group)
            mid = n // 2 + (n % 2) # median
            for i, (_, row) in enumerate(group.iterrows()):
                nisn_raw = str(row.get('NISN / NIS', '')).split('/')[0].strip()
                nisn_clean = nisn_raw.lstrip('0')
                if i < mid:
                    girl_group_map[nisn_clean] = f"GC1 {kelas_fmt}"
                else:
                    girl_group_map[nisn_clean] = f"GC2 {kelas_fmt}"

        try:
            df_form = pd.read_excel(input_file, sheet_name='Form Presensi SCC')
        except:
            df_form = pd.DataFrame(columns=['NISN', 'Pertemuan Grup', 'Status'])

        total_pertemuan = 7
        records = []
        excel_row = 2
        for index, row in df_master.iterrows():
            nama_mentah = str(row.get('NAMA', ''))
            nama_bersih = nama_mentah.split(' - ')[0] if ' - ' in nama_mentah else nama_mentah
            
            nisn_nis = str(row.get('NISN / NIS', ''))
            nisn = nisn_nis.split('/')[0].strip() if '/' in nisn_nis else nisn_nis
            nis = nisn_nis.split('/')[1].strip() if '/' in nisn_nis else ''
            
            angkatan = row.get('ANGKATAN', '')
            kelas = row.get('KELAS TA', '')
            kelompok = row.get('Grup TA', '')
            mentor = row.get('Mentor TA', '')
            lp = row.get('L/P', '')
            
            keterangan = ""
            status_warna = ""
            if str(angkatan) == '33':
                nisn_lookup = nisn.lstrip('0')
                kelas_str = str(kelas).strip()
                kelas_fmt = kelas_str.replace('X', 'X-') if kelas_str.startswith('X') and '-' not in kelas_str else kelas_str
                
                # Check SMAVO
                if nisn_lookup in mentor_map:
                    if mentor_map[nisn_lookup]['Grup']: kelompok = mentor_map[nisn_lookup]['Grup']
                    if mentor_map[nisn_lookup]['Mentor']: mentor = mentor_map[nisn_lookup]['Mentor']
                
                # AUTO-ASSIGN
                if (pd.isna(kelompok) or kelompok == "") and not is_nonis(nama_bersih):
                    if str(lp).strip().upper() == 'L':
                        kelompok = f"BC {kelas_fmt}"
                    elif str(lp).strip().upper() == 'P':
                        if nisn_lookup in girl_group_map:
                            kelompok = girl_group_map[nisn_lookup]
                        else:
                            kelompok = f"GC1 {kelas_fmt}"
                            
                    if kelompok in MASTER_MENTOR:
                        mentor = MASTER_MENTOR[kelompok]
                        
                # Pewarnaan & Keterangan HANYA untuk Angkatan 33
                if (pd.isna(kelompok) or kelompok == "") and (pd.isna(mentor) or mentor == ""):
                    if is_nonis(nama_bersih):
                        keterangan = "Nonis"
                        status_warna = "merah"
                    else:
                        keterangan = "Belum Terdaftar Kelompok"
                        status_warna = "hijau"
            
            p_status = {f"P{i}": "" for i in range(1, 8)}
            student_logs = df_form[df_form['NISN'].astype(str).str.contains(nisn, na=False)] if not df_form.empty and nisn else pd.DataFrame()
            hadir_count = 0
            for _, log_row in student_logs.iterrows():
                pertemuan = log_row.get('Pertemuan Grup')
                status = log_row.get('Status', '')
                if pd.notna(pertemuan) and str(pertemuan).isdigit():
                    p_idx = f"P{int(pertemuan)}"
                    if p_idx in p_status:
                        p_status[p_idx] = status
                        if str(status).lower() == 'hadir': hadir_count += 1
            
            if hadir_count == 0 and pd.notna(row.get('Total Kehadiran')):
                try: hadir_count = int(row.get('Total Kehadiran', 0))
                except: pass
            
            total_pertemuan = 7
            persentase = (hadir_count / total_pertemuan) * 100 if total_pertemuan > 0 else 0
            
            records.append({
                'No': index + 1,
                'NISN': nisn,
                'NIS': nis,
                'Nama peserta didik': nama_bersih,
                'L/P': lp,
                'total pertemuan': total_pertemuan,
                'total kehadiran': f'=COUNTIF(O{excel_row}:U{excel_row}; "*Hadir*")',
                '% kehadiran': f"=G{excel_row}/F{excel_row}",
                'nilai rapor': "",
                'Keterangan': keterangan,
                'Warna': status_warna,
                'Angkatan': angkatan,
                'Kelas': kelas,
                'Kelompok': (f"{kelompok} {int(angkatan)}" if pd.notna(kelompok) and str(kelompok).strip() and pd.notna(angkatan) and str(angkatan).isdigit() and not str(kelompok).strip().endswith(str(int(angkatan))) else (kelompok if pd.notna(kelompok) else "")),
                'nama mentor': mentor if pd.notna(mentor) else "",
                'P1': p_status['P1'],
                'P2': p_status['P2'],
                'P3': p_status['P3'],
                'P4': p_status['P4'],
                'P5': p_status['P5'],
                'P6': p_status['P6'],
                'P7': p_status['P7']
            })
            excel_row += 1
            
        df_output = pd.DataFrame(records)
        df_color = df_output.pop('Warna') 
        
        print(f"Menyimpan ke {output_file}...")
        df_output.to_excel(output_file, index=False, sheet_name='Rekap Presensi')
        
        wb = load_workbook(output_file)
        ws = wb['Rekap Presensi']
        
        pastel_red = PatternFill(start_color="FFD1DC", end_color="FFD1DC", fill_type="solid")
        pastel_green = PatternFill(start_color="D1FFD1", end_color="D1FFD1", fill_type="solid")
        
        for idx, row in enumerate(ws.iter_rows(min_row=2)):
            cell_warna = df_color.iloc[idx]
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
                if cell.column_letter == 'H':
                    cell.number_format = '0.0%'
                if cell_warna == "merah":
                    cell.fill = pastel_red
                elif cell_warna == "hijau":
                    cell.fill = pastel_green
                    
        wb.save(output_file)
        print("Selesai! File Excel berhasil dibuat.")
        
        # --- 5. Sync ke Google Sheets via GAS ---
        try:
            import gas_sync
            gas_sync.send_to_google_sheets(records, df_color)
        except Exception as e:
            print("Gagal memanggil gas_sync:", e)
            
    except Exception as e:
        print(f"Terjadi kesalahan: {e}")

if __name__ == "__main__":
    generate_rekap()
