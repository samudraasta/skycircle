import pandas as pd
from openpyxl import load_workbook

output_file = "Rekap_Presensi_SCC_X_XI_XII.xlsx"
df = pd.read_excel(output_file)

nonis_df = df[df['Keterangan'] == 'Nonis']
print("Number of students marked as Nonis in Pandas DataFrame:", len(nonis_df))
if len(nonis_df) > 0:
    print(nonis_df[['NISN', 'Nama peserta didik', 'Keterangan']].head())

# Check openpyxl directly
wb = load_workbook(output_file)
ws = wb['Rekap Presensi']

headers = {cell.value: i for i, cell in enumerate(ws[1])}
ket_col = headers.get('Keterangan')

red_count = 0
ket_count = 0
if ket_col is not None:
    for row in ws.iter_rows(min_row=2):
        cell = row[ket_col]
        if cell.value == 'Nonis':
            ket_count += 1
        
        # Check fill color of the first cell
        if row[0].fill.start_color.index != '00000000': # '00000000' is usually transparent
            color = row[0].fill.start_color.index
            if color == '00FFD1DC': # ARGB format for pastel_red
                red_count += 1

print(f"Number of rows with 'Nonis' in Excel cell: {ket_count}")
print(f"Number of rows with pastel_red fill: {red_count}")
