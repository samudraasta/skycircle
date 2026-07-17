from openpyxl import load_workbook

wb = load_workbook("Rekap_Presensi_SCC_X_XI_XII.xlsx")
ws = wb.active

print("Cell H2 value:", repr(ws['H2'].value))
print("Cell H2 data type:", ws['H2'].data_type)
